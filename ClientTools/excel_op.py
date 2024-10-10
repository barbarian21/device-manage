
from openpyxl import Workbook
from openpyxl.styles import Border,Side,Font,colors,Alignment,numbers
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image

def textToDigit2(arg):
    if arg.replace('.','').replace('-','').replace(',','').isdigit():
        return float(arg.replace(',',''))
    else:
        return arg  

def create_excel(results,ws,cards,pathRes):
    #pics = draw_pic(results[2])
    # border = Border(
    #     left=Side(border_style='thin',color='000000'),
    #     right=Side(border_style='thin',color='000000'),
    #     top=Side(border_style='thin',color='000000'),
    #     bottom=Side(border_style='thin',color='000000')
    #     )
    wb = Workbook()
    wb.guess_types=True
    ws0 = wb.worksheets[0]
    ws0.title = 'cse-servers'
    wb.create_sheet('cse-workstations')
    ws1 = wb.worksheets[1]
    wb.create_sheet('workstation')
    ws2 = wb.worksheets[2]

    # imgsize = (1280 / 4, 1280 / 4)  # 设置一个图像缩小的比例
    # ws0.column_dimensions['A'].width = imgsize[0] * 0.14  # 修改列A的宽
    # img = Image(pics[0])  # 缩放图片
    # img.width, img.height = imgsize
    # ws0.add_image(img, 'A24')  # 图片 插入 A1 的位置上
    # ws0.row_dimensions[1].height = imgsize[1] * 0.78 
    for item in results[1]:
        ws0.append(item)

    for item in results[0]:
        ws1.append(item)
    for item in ws:
        ws2.append(item)

    # write cards to different sheet
    for item in cards:
        sheet_name = 'cse-' + item
        wb.create_sheet(sheet_name)
        w_s = wb[sheet_name]
        for line in cards[item]:
            w_s.append(line)

    # endIndex0 = 'H'+str(len(results[1]))
    # endIndex1 = 'H'+str(len(results[0]))
    # cell0s = ws0['A1':endIndex0]
    # cell1s = ws1['A1':endIndex1]

    # endIndex2 = 'H' + str(len(ws))
    # cell2s = ws2['A1':endIndex2]

    # for row in cell0s:
    #     for cell in row:
    #         cell.border = border
    # for row in cell1s:
    #     for cell in row:
    #         cell.border = border

    # for row in cell2s:
    #     for cell in row:
    #         cell.border = border
    wb.save(pathRes)
